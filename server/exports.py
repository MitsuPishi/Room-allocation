"""Generate authenticated run artifacts for university operations."""

from __future__ import annotations

import hashlib
from html import escape
import json
import logging
from pathlib import Path

from jinja2 import BaseLoader, Environment
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from .config import settings
from .storage import run_artifact_dir


logger = logging.getLogger(__name__)


PERSIAN_HEADERS = {
    "room_id": "شناسه اتاق",
    "bed": "تخت",
    "room_capacity": "ظرفیت اتاق",
    "room_size": "تعداد ساکنان",
    "student_idx": "ردیف دانشجو",
    "student_id": "شناسه دانشجو",
    "student_name": "نام دانشجو",
    "student_utility": "امتیاز دانشجو",
    "room_quality": "کیفیت اتاق",
    "mean_student_utility": "میانگین امتیاز",
    "faculty": "دانشکده",
    "major": "رشته",
    "age": "سن",
    "sleep_window": "بازه خواب",
    "wake_window": "بازه بیداری",
    "noise_tolerance": "تحمل صدا",
    "study_habit": "عادت مطالعه",
    "cleanliness": "نظافت",
    "severity": "سطح",
    "code": "کد",
    "row": "ردیف",
    "field": "فیلد",
    "value": "مقدار",
    "message": "پیام",
    "cleanliness_contribution": "سهم نظافت",
    "noise_contribution": "سهم تحمل صدا",
    "study_contribution": "سهم مطالعه",
    "schedule_contribution": "سهم برنامه خواب",
}


def _json_default(value: object) -> object:
    if hasattr(value, "item"):
        return value.item()
    return str(value)


def _safe(value: object) -> object:
    if pd.isna(value):
        return ""
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        # Prevent spreadsheet formula execution from questionnaire-controlled text.
        return "'" + value
    return value


def _style_sheet(sheet, *, freeze: str = "A2") -> None:
    sheet.sheet_view.rightToLeft = True
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = freeze
    header_fill = PatternFill("solid", fgColor="173F3B")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    subtle = Side(style="thin", color="D9E4E2")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=subtle)
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        values = [str(cell.value or "") for cell in column[:200]]
        width = min(42, max(11, max((len(value) for value in values), default=8) + 2))
        sheet.column_dimensions[get_column_letter(column[0].column)].width = width
        for cell in column[1:]:
            cell.alignment = Alignment(horizontal="right", vertical="center")


def _write_frame(workbook: Workbook, title: str, frame: pd.DataFrame) -> None:
    sheet = workbook.create_sheet(title)
    headers = [PERSIAN_HEADERS.get(column, column) for column in frame.columns]
    sheet.append(headers)
    for row in frame.itertuples(index=False, name=None):
        sheet.append([_safe(value) for value in row])
    _style_sheet(sheet)


def _write_summary(workbook: Workbook, metadata: dict) -> None:
    sheet = workbook.active
    sheet.title = "خلاصه"
    sheet.sheet_view.rightToLeft = True
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:D2")
    sheet["A1"] = "گزارش تخصیص اتاق یونی‌میت"
    sheet["A1"].fill = PatternFill("solid", fgColor="173F3B")
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    metrics = metadata.get("metrics", {})
    inventory = metadata.get("room_inventory", {})
    rows = [
        ("وضعیت", metadata.get("status", "")),
        ("نسخه الگوریتم", metadata.get("algorithm_version", "")),
        ("تعداد دانشجویان", inventory.get("occupied_beds", 0)),
        ("اتاق‌های فعال", inventory.get("assigned_rooms", 0)),
        ("اتاق‌های استفاده‌نشده", inventory.get("unused_rooms", 0)),
        ("تخت خالی در اتاق‌های فعال", inventory.get("active_vacancies", 0)),
        ("کمترین امتیاز دانشجو", metrics.get("min_student_utility", 0)),
        ("صدک دهم کیفیت اتاق", metrics.get("p10_room_quality", 0)),
        ("میانگین امتیاز دانشجویان", metrics.get("mean_student_utility", 0)),
        ("زمان اجرا (ثانیه)", metadata.get("runtime_seconds", 0)),
    ]
    for row_index, (label, value) in enumerate(rows, start=4):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)
        sheet.cell(row=row_index, column=1).font = Font(bold=True, color="173F3B")
        sheet.cell(row=row_index, column=1).fill = PatternFill("solid", fgColor="EAF2F0")
        sheet.cell(row=row_index, column=1).alignment = Alignment(horizontal="right")
        sheet.cell(row=row_index, column=2).alignment = Alignment(horizontal="right")
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 24


PDF_TEMPLATE = """
<!doctype html>
<html lang="fa" dir="rtl">
<head><meta charset="utf-8"><style>
@page { size: A4; margin: 14mm; @bottom-center { content: "صفحه " counter(page); } }
body { font-family: "DejaVu Sans", sans-serif; color: #173f3b; font-size: 10px; }
h1 { background: #173f3b; color: white; padding: 16px; border-radius: 10px; font-size: 21px; }
h2 { color: #a76532; margin-top: 22px; font-size: 15px; }
.cards { display: flex; flex-wrap: wrap; gap: 8px; }
.card { width: 21%; background: #edf4f2; padding: 10px; border-right: 4px solid #2f7f75; }
.card strong { display: block; font-size: 16px; margin-top: 5px; }
table { border-collapse: collapse; width: 100%; margin-top: 8px; page-break-inside: auto; }
th { background: #173f3b; color: white; padding: 7px; }
td { border-bottom: 1px solid #d9e4e2; padding: 6px; text-align: right; }
tr { page-break-inside: avoid; }
.muted { color: #607b77; }
</style></head>
<body>
<h1>گزارش رسمی تخصیص اتاق یونی‌میت</h1>
<p class="muted">شناسه اجرا: {{ run_id }}</p>
<div class="cards">
  <div class="card">دانشجویان<strong>{{ inventory.occupied_beds }}</strong></div>
  <div class="card">اتاق‌های فعال<strong>{{ inventory.assigned_rooms }}</strong></div>
  <div class="card">کمترین امتیاز<strong>{{ "%.1f"|format(metrics.min_student_utility) }}</strong></div>
  <div class="card">میانگین امتیاز<strong>{{ "%.1f"|format(metrics.mean_student_utility) }}</strong></div>
</div>
<h2>خلاصه موجودی</h2>
<table><tr><th>کل اتاق‌ها</th><th>اتاق استفاده‌نشده</th><th>تخت فعال</th><th>تخت خالی فعال</th></tr>
<tr><td>{{ inventory.total_rooms }}</td><td>{{ inventory.unused_rooms }}</td><td>{{ inventory.active_beds }}</td><td>{{ inventory.active_vacancies }}</td></tr></table>
<h2>فهرست اتاق‌ها و ساکنان</h2>
{% for room in rooms %}
<h3>{{ room.room_id }} - ظرفیت {{ room.room_capacity }} - کیفیت {{ "%.1f"|format(room.room_quality) }}</h3>
<table><tr><th>تخت</th><th>شناسه دانشجو</th><th>نام دانشجو</th><th>امتیاز</th></tr>
{% for student in room.students %}<tr><td>{{ student.bed }}</td><td>{{ student.student_id }}</td><td>{{ student.student_name }}</td><td>{{ "%.1f"|format(student.student_utility) }}</td></tr>{% endfor %}
</table>
{% endfor %}
</body></html>
"""


def _write_pdf(html: str, output_dir: Path, pdf_path: Path) -> None:
    from weasyprint import HTML

    HTML(string=html, base_url=str(output_dir)).write_pdf(pdf_path)


def generate_run_artifacts(
    run_id: str,
    assignments: pd.DataFrame,
    rooms: pd.DataFrame,
    students: pd.DataFrame,
    validation: pd.DataFrame,
    metadata: dict,
) -> dict[str, dict[str, object]]:
    output_dir = run_artifact_dir(run_id)
    files: dict[str, Path] = {}

    csv_frames = {
        "assignments.csv": assignments,
        "room_metrics.csv": rooms,
        "student_metrics.csv": students,
        "validation_report.csv": validation,
    }
    for name, frame in csv_frames.items():
        path = output_dir / name
        safe_frame = frame.copy()
        for column in safe_frame.select_dtypes(include=["object", "string"]).columns:
            safe_frame[column] = safe_frame[column].map(_safe)
        safe_frame.to_csv(path, index=False, encoding="utf-8-sig")
        files[name] = path

    metadata_path = output_dir / "run_metadata.json"
    metadata_path.write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2, default=_json_default),
        encoding="utf-8",
    )
    files[metadata_path.name] = metadata_path

    workbook = Workbook()
    _write_summary(workbook, metadata)
    _write_frame(workbook, "تخصیص‌ها", assignments)
    _write_frame(workbook, "اتاق‌ها", rooms)
    _write_frame(workbook, "دانشجویان", students)
    _write_frame(workbook, "اعتبارسنجی", validation)
    workbook_path = output_dir / "unimate_report.xlsx"
    workbook.save(workbook_path)
    files[workbook_path.name] = workbook_path

    room_records = []
    for room in rooms.to_dict("records"):
        roster = assignments.loc[assignments["room_id"] == room["room_id"]]
        room_records.append(
            {
                **room,
                "students": roster[
                    ["bed", "student_id", "student_name", "student_utility"]
                ].to_dict("records"),
            }
        )
    html = Environment(loader=BaseLoader(), autoescape=True).from_string(
        PDF_TEMPLATE
    ).render(
        run_id=escape(run_id),
        inventory=metadata["room_inventory"],
        metrics=metadata["metrics"],
        rooms=room_records,
    )
    pdf_path = output_dir / "unimate_report.pdf"
    try:
        _write_pdf(html, output_dir, pdf_path)
    except (ImportError, OSError) as exc:
        if settings.environment == "production":
            raise
        logger.warning(
            "PDF export is unavailable in this development environment (%s).",
            type(exc).__name__,
        )
    else:
        files[pdf_path.name] = pdf_path

    manifest: dict[str, dict[str, object]] = {}
    for name, path in files.items():
        content = path.read_bytes()
        manifest[name] = {
            "storage_key": str(path.relative_to(output_dir.parent.parent)),
            "sha256": hashlib.sha256(content).hexdigest(),
            "size": len(content),
        }
    return manifest
