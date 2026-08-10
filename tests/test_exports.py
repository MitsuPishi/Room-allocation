import hashlib
import sys
from types import SimpleNamespace

from openpyxl import load_workbook
import pandas as pd

from server import exports


def test_artifact_bundle_has_stable_schemas_rtl_workbook_and_hashes(tmp_path, monkeypatch):
    output_dir = tmp_path / "artifacts" / "run-1"

    def artifact_dir(run_id: str):
        target = tmp_path / "artifacts" / run_id
        target.mkdir(parents=True, exist_ok=True)
        return target

    class FakeHTML:
        def __init__(self, *, string: str, base_url: str):
            assert "گزارش رسمی تخصیص اتاق" in string
            assert base_url == str(output_dir)

        def write_pdf(self, path):
            path.write_bytes(b"%PDF-1.7\nartifact-test")

    monkeypatch.setattr(exports, "run_artifact_dir", artifact_dir)
    monkeypatch.setitem(sys.modules, "weasyprint", SimpleNamespace(HTML=FakeHTML))

    assignments = pd.DataFrame(
        [
            {
                "room_id": "B-401",
                "bed": 1,
                "room_capacity": 4,
                "student_idx": 0,
                "student_id": "S-001",
                "student_name": "=HYPERLINK(\"https://invalid.example\")",
                "student_utility": 92.5,
                "room_quality": 91.0,
            }
        ]
    )
    rooms = pd.DataFrame(
        [
            {
                "room_id": "B-401",
                "room_size": 1,
                "room_capacity": 4,
                "room_quality": 91.0,
                "mean_student_utility": 92.5,
            }
        ]
    )
    students = assignments[
        ["student_idx", "student_id", "room_id", "student_utility"]
    ]
    validation = pd.DataFrame(
        columns=["severity", "code", "row", "field", "value", "message"]
    )
    metadata = {
        "status": "succeeded",
        "algorithm_version": "test",
        "runtime_seconds": 0.1,
        "metrics": {
            "min_student_utility": 92.5,
            "p10_room_quality": 91.0,
            "mean_student_utility": 92.5,
        },
        "room_inventory": {
            "occupied_beds": 1,
            "assigned_rooms": 1,
            "unused_rooms": 1,
            "active_vacancies": 3,
            "total_rooms": 2,
            "active_beds": 4,
        },
    }

    manifest = exports.generate_run_artifacts(
        "run-1", assignments, rooms, students, validation, metadata
    )

    assert {
        "assignments.csv",
        "room_metrics.csv",
        "student_metrics.csv",
        "validation_report.csv",
        "run_metadata.json",
        "unimate_report.xlsx",
        "unimate_report.pdf",
    } == set(manifest)
    for item in manifest.values():
        artifact = tmp_path / item["storage_key"]
        content = artifact.read_bytes()
        assert hashlib.sha256(content).hexdigest() == item["sha256"]
        assert len(content) == item["size"]

    csv_text = (output_dir / "assignments.csv").read_text(encoding="utf-8-sig")
    assert csv_text.splitlines()[0].startswith(
        "room_id,bed,room_capacity,student_idx,student_id"
    )
    assert "'=HYPERLINK" in csv_text

    workbook = load_workbook(output_dir / "unimate_report.xlsx")
    assert workbook.sheetnames == [
        "خلاصه",
        "تخصیص‌ها",
        "اتاق‌ها",
        "دانشجویان",
        "اعتبارسنجی",
    ]
    assert all(sheet.sheet_view.rightToLeft for sheet in workbook.worksheets)
    assignment_sheet = workbook["تخصیص‌ها"]
    assert assignment_sheet["A1"].value == "شناسه اتاق"
    assert assignment_sheet["F2"].value.startswith("'=")

    def unavailable_pdf(*args, **kwargs):
        del args, kwargs
        raise OSError("native PDF libraries unavailable")

    monkeypatch.setattr(exports, "_write_pdf", unavailable_pdf)
    without_pdf = exports.generate_run_artifacts(
        "run-2", assignments, rooms, students, validation, metadata
    )
    assert "unimate_report.pdf" not in without_pdf
    assert "unimate_report.xlsx" in without_pdf
